"""Lectura del Excel de hojas de vida (HOJAS_DE_VIDA.xlsm).

El archivo tiene una hoja por equipo -- 963 en la version de agosto de 2026 --
mas una hoja indice que dice que equipo corresponde a cada una. Se lee el
indice primero y solo despues se abren las hojas que el indice nombra. El
prototipo abria las 963 antes de mirar el indice, y cargaba el libro entero
en memoria para usar unas 750.

En el PC servidor la ruta apunta al archivo original de la unidad de red, que
la unidad tiene abierto mientras el ETL corre. Se abre en modo solo lectura,
que no bloquea a quien lo este editando.
"""

import logging
import shutil
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from core.config import obtener_ajustes
from etl.transform.normalizar import es_nic_valido, normalizar_clave, normalizar_serie

# La hoja indice conserva su nombre de 2019 aunque ya no describa solo
# equipos criticos. Las planillas no se tocan: el ETL se adapta.
HOJA_INDICE = "EQUIPOS CRITICOS 2019"

# En la hoja indice el encabezado esta en la fila 6; arriba va el titulo.
FILA_ENCABEZADO_INDICE = 6

# En las hojas de equipo el encabezado no siempre cae en la misma fila:
# se busca entre estas, que es donde aparece en la practica.
FILAS_POSIBLES_DE_ENCABEZADO = range(5, 13)

_COLUMNAS_OBLIGATORIAS_DEL_INDICE = ("EQUIPO", "NIC", "SERIE")

_registro = logging.getLogger(__name__)


class EstructuraInesperada(RuntimeError):
    """El archivo no tiene la forma que el ETL espera."""


@dataclass(frozen=True)
class EquipoDelIndice:
    """Una fila del indice: a que equipo pertenece cada hoja del libro."""

    equipo: str | None
    nic: str | None
    serie: str | None
    servicio: str | None
    marca: str | None
    modelo: str | None

    @property
    def clave_de_hoja(self) -> str:
        """Con que nombre buscar su hoja: primero el NIC, si no la serie."""
        return normalizar_clave(self.nic) or normalizar_clave(self.serie)


def abrir_libro(ruta: Path | None = None) -> Any:
    """Abre el .xlsm en solo lectura, sin bloquear a quien lo este editando.

    Si la lectura falla porque alguien esta guardando en ese momento, se
    trabaja sobre una copia y queda anotado en el registro, junto con la hora
    del archivo, para saber a que momento corresponden los datos cargados.
    """
    archivo = ruta or obtener_ajustes().exigir_excel_hoja_de_vida()

    try:
        return load_workbook(archivo, read_only=True, data_only=True, keep_vba=False)
    except (OSError, PermissionError) as error:
        _registro.warning(
            "No se pudo abrir %s directamente (%s). Se leera una copia.", archivo, error
        )
        copia = _copiar_a_temporal(archivo)
        return load_workbook(copia, read_only=True, data_only=True, keep_vba=False)


def _copiar_a_temporal(archivo: Path) -> Path:
    destino = Path(tempfile.gettempdir()) / f"hrc_cems_{archivo.name}"
    shutil.copy2(archivo, destino)
    modificado = datetime.fromtimestamp(archivo.stat().st_mtime)
    _registro.warning(
        "Copia leida desde %s. El original fue modificado por ultima vez el %s.",
        destino,
        modificado.strftime("%d-%m-%Y %H:%M"),
    )
    return destino


def leer_indice(libro: Any) -> list[EquipoDelIndice]:
    """Lee la hoja indice: la lista de equipos con hoja de vida propia."""
    if HOJA_INDICE not in libro.sheetnames:
        raise EstructuraInesperada(
            f"El libro no tiene la hoja indice '{HOJA_INDICE}'. "
            f"Hojas encontradas: {len(libro.sheetnames)}"
        )

    hoja = libro[HOJA_INDICE]
    filas = hoja.iter_rows(min_row=FILA_ENCABEZADO_INDICE, values_only=True)

    encabezado = next(filas, None)
    if encabezado is None:
        raise EstructuraInesperada(f"La hoja '{HOJA_INDICE}' esta vacia")

    posiciones = _ubicar_columnas(encabezado, _COLUMNAS_OBLIGATORIAS_DEL_INDICE)

    equipos: list[EquipoDelIndice] = []
    for fila in filas:
        nic = _celda(fila, posiciones.get("NIC"))
        serie = _celda(fila, posiciones.get("SERIE"))
        if not es_nic_valido(nic) and not serie:
            continue
        equipos.append(
            EquipoDelIndice(
                equipo=_celda(fila, posiciones.get("EQUIPO")),
                nic=nic if es_nic_valido(nic) else None,
                serie=normalizar_serie(serie),
                servicio=_celda(fila, posiciones.get("SERVICIO")),
                marca=_celda(fila, posiciones.get("MARCA")),
                modelo=_celda(fila, posiciones.get("MODELO")),
            )
        )

    _registro.info("Indice leido: %s equipos con hoja de vida", len(equipos))
    return equipos


def leer_intervenciones(libro: Any, nombre_hoja: str) -> list[dict[str, Any]]:
    """Lee las intervenciones registradas en la hoja de un equipo.

    Devuelve los valores como vienen: la fecha llega como la guardo Excel y el
    tipo como la marca `X` que puso la unidad. Interpretarlos es tarea de la
    capa de transformacion.
    """
    if nombre_hoja not in libro.sheetnames:
        return []

    filas = list(libro[nombre_hoja].iter_rows(values_only=True))
    posicion_encabezado = _buscar_encabezado(filas)
    if posicion_encabezado is None:
        _registro.debug("La hoja '%s' no tiene una tabla de intervenciones", nombre_hoja)
        return []

    encabezado = filas[posicion_encabezado]
    posiciones = _ubicar_columnas(encabezado, ("FECHA",))
    for opcional in ("MP", "MC", "ACTIVIDAD", "RESPALDO", "GARANTIA", "GARANTÍA"):
        posiciones.setdefault(opcional, _indice_de(encabezado, opcional))

    intervenciones: list[dict[str, Any]] = []
    for fila in filas[posicion_encabezado + 1 :]:
        fecha = fila[posiciones["FECHA"]] if posiciones.get("FECHA") is not None else None
        actividad = _celda(fila, posiciones.get("ACTIVIDAD"))
        if fecha is None and not actividad:
            continue
        intervenciones.append(
            {
                "hoja": nombre_hoja,
                "fecha": fecha,
                "preventivo": _marcado(fila, posiciones.get("MP")),
                "correctivo": _marcado(fila, posiciones.get("MC")),
                "actividad": actividad,
                "respaldo": _celda(fila, posiciones.get("RESPALDO")),
                "garantia": _celda(fila, posiciones.get("GARANTIA"))
                or _celda(fila, posiciones.get("GARANTÍA")),
            }
        )

    return intervenciones


# ----------------------------------------------------------------------
# Ayudas internas
# ----------------------------------------------------------------------


def _buscar_encabezado(filas: list[tuple[Any, ...]]) -> int | None:
    """Ubica la fila de encabezado de la tabla de intervenciones.

    La reconoce por tener una columna FECHA y las marcas MP o MC, que son las
    que distinguen mantenimiento preventivo de correctivo.
    """
    for posicion in FILAS_POSIBLES_DE_ENCABEZADO:
        if posicion >= len(filas):
            break
        titulos = {normalizar_clave(celda) for celda in filas[posicion] if celda is not None}
        if "FECHA" in titulos and ({"MP", "MC"} & titulos):
            return posicion
    return None


def _ubicar_columnas(
    encabezado: tuple[Any, ...], obligatorias: tuple[str, ...]
) -> dict[str, int | None]:
    posiciones: dict[str, int | None] = {}
    for columna in obligatorias:
        posicion = _indice_de(encabezado, columna)
        if posicion is None:
            raise EstructuraInesperada(
                f"Falta la columna '{columna}'. Encabezado leido: "
                f"{[c for c in encabezado[:15] if c]}"
            )
        posiciones[columna] = posicion
    for columna in ("SERVICIO", "MARCA", "MODELO"):
        posiciones.setdefault(columna, _indice_de(encabezado, columna))
    return posiciones


def _indice_de(encabezado: tuple[Any, ...], columna: str) -> int | None:
    buscada = normalizar_clave(columna)
    for posicion, celda in enumerate(encabezado):
        if celda is not None and normalizar_clave(celda) == buscada:
            return posicion
    return None


def _celda(fila: tuple[Any, ...], posicion: int | None) -> str | None:
    if posicion is None or posicion >= len(fila):
        return None
    valor = fila[posicion]
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def _marcado(fila: tuple[Any, ...], posicion: int | None) -> bool:
    """La unidad marca con una X la columna que corresponde."""
    valor = _celda(fila, posicion)
    return bool(valor) and valor.strip().upper() == "X"
